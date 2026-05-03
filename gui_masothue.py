import sys
sys.stdout.reconfigure(encoding='utf-8')

import tkinter as tk
from tkinter import filedialog, messagebox
import threading
import time
import os
import types
import re

# ==========================================
# 1. POLYFILL VÀ CÁC HÀM CÀO DỮ LIỆU CỐT LÕI
# ==========================================

# Polyfill phải nằm TRƯỚC KHI import undetected_chromedriver
if 'distutils' not in sys.modules:
    class LooseVersion:
        def __init__(self, vstring):
            self.vstring = str(vstring)
            self.version = [int(x) if x.isdigit() else x for x in re.split(r'([0-9]+|\.)', self.vstring) if x and x != '.']
        def __str__(self): return self.vstring
        def __ge__(self, other): return True
        def __lt__(self, other): return False

    d = types.ModuleType('distutils')
    d.version = types.ModuleType('distutils.version')
    d.version.LooseVersion = LooseVersion
    sys.modules['distutils'] = d
    sys.modules['distutils.version'] = d.version

import pandas as pd
import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

def format_excel(file_path):
    """Định dạng file Excel để trông chuyên nghiệp hơn"""
    try:
        # Đợi một chút để đảm bảo file đã được đóng hoàn toàn từ pandas
        time.sleep(1)
        wb = load_workbook(file_path)
        ws = wb.active
        
        # Định nghĩa các style
        # Font mặc định: Times New Roman, size 11
        default_font = Font(name='Times New Roman', size=11)
        header_font = Font(name='Times New Roman', bold=True, size=11)
        center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        
        # 1. Áp dụng font, border và căn lề cho TOÀN BỘ các ô có dữ liệu
        for row in ws.iter_rows():
            for cell in row:
                cell.font = default_font
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center", wrap_text=True)

        # 2. Định dạng riêng cho dòng tiêu đề (Headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = center_alignment
            # Bỏ tô màu nền
            cell.fill = PatternFill(fill_type=None)
            
        # 3. Cố định dòng tiêu đề (Freeze Panes)
        ws.freeze_panes = "A2"
        
        # 4. Tạo bộ lọc Filter cho vùng dữ liệu
        ws.auto_filter.ref = ws.dimensions
                
        # 5. Tự động điều chỉnh độ rộng cột
        for idx, column in enumerate(ws.columns):
            max_length = 0
            column_letter = column[0].column_letter
            
            # Duyệt qua các ô trong cột để tìm độ dài lớn nhất
            for cell in column:
                try:
                    if cell.value:
                        lines = str(cell.value).split('\n')
                        line_max = max([len(line) for line in lines])
                        if line_max > max_length:
                            max_length = line_max
                except:
                    pass
            
            # Căn chỉnh kỹ cho 4 cột đầu tiên
            if idx < 4:
                adjusted_width = max(15, min(max_length + 5, 50))
            else:
                adjusted_width = max(10, min(max_length + 2, 60))
                
            ws.column_dimensions[column_letter].width = adjusted_width
            
        wb.save(file_path)
        print(f" -> Đã định dạng Excel: Times New Roman, Filter, Freeze Panes.")
    except Exception as e:
        print(f" -> Cảnh báo: Không thể định dạng file Excel. Chi tiết: {e}")

def setup_driver():
    options = uc.ChromeOptions()
    options.add_argument('--no-first-run --no-service-autorun --password-store=basic')
    driver = uc.Chrome(options=options, version_main=147)
    return driver

def extract_data_from_html(html, tax_code):
    soup = BeautifulSoup(html, 'html.parser')
    data = {}
    
    company_name = "Không tìm thấy"
    name_tag = soup.find('th', itemprop='name') or soup.find('span', itemprop='name')
    if name_tag:
        company_name = name_tag.text.strip()
    else:
        h1 = soup.find('h1')
        if h1: company_name = h1.text.strip()
    data['Tên công ty'] = company_name

    table = soup.find('table', class_='table-taxinfo')
    if table:
        for row in table.find_all('tr'):
            tds = row.find_all(['td', 'th'])
            if len(tds) == 2:
                label = tds[0].text.strip()
                value = tds[1].text.strip()
                if 'Ngành nghề' in label or 'Mã ngành' in label:
                    label = 'Ngành nghề chính'
                    a_tag = tds[1].find('a')
                    if a_tag:
                        value = a_tag.text.strip()
                if label and label != company_name:
                    data[label] = value
                
    all_industries = []
    industry_tables = soup.find_all('table', class_='table')
    target_table = None
    
    for tbl in industry_tables:
        header = tbl.find('th')
        if header and ('Mã' in header.text or 'Ngành' in header.text):
            target_table = tbl
            break
            
    if not target_table and industry_tables:
        target_table = industry_tables[0]
        
    if target_table:
        for row in target_table.find_all('tr'):
            tds = row.find_all('td')
            if len(tds) >= 2:
                code = tds[0].text.strip()
                name = " ".join(tds[1].text.split())
                is_main = False
                if 'chính' in row.text.lower() or 'strong' in str(row):
                    is_main = True
                    if 'Ngành nghề chính' not in data or data['Ngành nghề chính'] == "":
                        data['Ngành nghề chính'] = f"{code} - {name}"
                    data['Mã Ngành Chính'] = code
                
                if is_main:
                    all_industries.append(f"[{code}] {name} (Ngành chính)")
                else:
                    all_industries.append(f"[{code}] {name}")
                    
    if all_industries:
        data['Danh sách ngành nghề'] = "\n".join(all_industries)
        
    # Nếu không lấy được mã ngành chính từ bảng chi tiết, dùng regex để bóc tách từ Ngành nghề chính
    if 'Mã Ngành Chính' not in data and 'Ngành nghề chính' in data:
        match = re.search(r'^\[?(\d{2,5})\]?[\s-]*', data['Ngành nghề chính'])
        if match:
            data['Mã Ngành Chính'] = match.group(1)
        else:
            data['Mã Ngành Chính'] = ""

    return data

# ==========================================
# 2. GIAO DIỆN PHẦN MỀM (GUI)
# ==========================================

class TextRedirector(object):
    def __init__(self, widget, tag="stdout"):
        self.widget = widget
        self.tag = tag

    def write(self, str):
        self.widget.configure(state="normal")
        self.widget.insert("end", str, (self.tag,))
        self.widget.see("end")
        self.widget.configure(state="disabled")
        self.widget.update()
        
    def flush(self):
        pass

class MaSoThueApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Công cụ trích xuất dữ liệu Mã Số Thuế")
        self.root.geometry("700x500")
        self.root.configure(padx=20, pady=20)
        
        self.input_file = ""
        self.is_running = False
        
        # Phần chọn file
        frame_top = tk.Frame(root)
        tk.Label(frame_top, text="Chọn file Excel chứa danh sách Mã số thuế:", font=("Arial", 11, "bold")).pack(anchor="w")
        
        self.lbl_file = tk.Label(frame_top, text="(Chưa chọn file)", fg="blue", font=("Arial", 10), wraplength=500, justify="left")
        self.lbl_file.pack(side=tk.LEFT, pady=10)
        
        tk.Button(frame_top, text="Duyệt File...", command=self.select_file, bg="#f0f0f0", font=("Arial", 10)).pack(side=tk.RIGHT)
        frame_top.pack(fill=tk.X, pady=(0, 10))
        
        # Cột mã số thuế
        frame_col = tk.Frame(root)
        tk.Label(frame_col, text="Tên cột chứa Mã số thuế trong Excel:", font=("Arial", 10)).pack(side=tk.LEFT)
        self.txt_col = tk.Entry(frame_col, width=20, font=("Arial", 10))
        self.txt_col.insert(0, "Mã số thuế")
        self.txt_col.pack(side=tk.LEFT, padx=10)
        frame_col.pack(fill=tk.X, pady=(0, 20))
        
        # Tùy chọn nâng cao
        frame_opts = tk.Frame(root)
        self.var_auto_run = tk.BooleanVar(value=True)
        self.var_auto_open = tk.BooleanVar(value=True)
        
        tk.Checkbutton(frame_opts, text="Tự động trích xuất ngay khi chọn file", variable=self.var_auto_run, font=("Arial", 10)).pack(side=tk.LEFT, padx=(0, 20))
        tk.Checkbutton(frame_opts, text="Tự động mở file kết quả khi hoàn tất", variable=self.var_auto_open, font=("Arial", 10)).pack(side=tk.LEFT)
        frame_opts.pack(fill=tk.X, pady=(0, 15))
        
        # Nút chạy
        self.btn_run = tk.Button(root, text="🚀 BẮT ĐẦU TRÍCH XUẤT", command=self.start_processing, bg="#4CAF50", fg="white", font=("Arial", 12, "bold"), height=2)
        self.btn_run.pack(fill=tk.X, pady=(0, 20))
        
        # Cửa sổ log
        tk.Label(root, text="Tiến trình hoạt động:", font=("Arial", 10, "bold")).pack(anchor="w")
        self.text_log = tk.Text(root, wrap="word", height=15, state="disabled", font=("Consolas", 9), bg="#f5f5f5")
        self.text_log.pack(fill=tk.BOTH, expand=True)
        
        # Redirect print() to text_log
        sys.stdout = TextRedirector(self.text_log, "stdout")

    def select_file(self):
        filepath = filedialog.askopenfilename(
            title="Chọn file danh sách",
            filetypes=(("Excel files", "*.xlsx *.xls"), ("All files", "*.*"))
        )
        if filepath:
            self.input_file = filepath
            self.lbl_file.config(text=filepath)
            if self.var_auto_run.get():
                self.start_processing()
            
    def start_processing(self):
        if not self.input_file:
            messagebox.showwarning("Cảnh báo", "Vui lòng chọn file Excel trước khi bắt đầu!")
            return
            
        col_name = self.txt_col.get().strip()
        if not col_name:
            messagebox.showwarning("Cảnh báo", "Vui lòng nhập tên cột chứa Mã số thuế!")
            return
            
        if self.is_running:
            return
            
        self.is_running = True
        self.btn_run.config(state="disabled", text="⏳ ĐANG XỬ LÝ... (VUI LÒNG KHÔNG ĐÓNG APP)")
        
        # Chạy trong luồng phụ để không bị đơ giao diện
        threading.Thread(target=self.run_scraper, args=(self.input_file, col_name), daemon=True).start()

    def run_scraper(self, input_file, col_name):
        try:
            output_file = input_file.replace(".xlsx", "_ket_qua.xlsx")
            print(f"Bắt đầu đọc file: {os.path.basename(input_file)}")
            
            try:
                df = pd.read_excel(input_file, dtype={col_name: str})
            except Exception as e:
                print(f"LỖI: Không thể đọc file Excel. Chi tiết: {e}")
                self.finish_processing()
                return

            if 'Tên công ty' not in df.columns: df['Tên công ty'] = ""
            if 'Mã Ngành Chính' not in df.columns: df['Mã Ngành Chính'] = ""
            if 'Ngành nghề chính' not in df.columns: df['Ngành nghề chính'] = ""
            if 'Trạng thái' not in df.columns: df['Trạng thái'] = ""

            print(f"Đã tải {len(df)} dòng dữ liệu. Khởi động trình duyệt giả lập...")
            try:
                driver = setup_driver()
            except Exception as e:
                print(f"LỖI: Không thể khởi động Chrome. Kiểm tra lại trình duyệt Chrome trên máy. Chi tiết: {e}")
                self.finish_processing()
                return
                
            for index, row in df.iterrows():
                code = str(row.get(col_name, "")).strip()
                if not code or code == 'nan' or row.get('Trạng thái') == 'Thành công':
                    continue
                    
                print(f"[{index+1}/{len(df)}] Đang lấy dữ liệu MST: {code}")
                
                try:
                    driver.get('https://masothue.com/')
                    time.sleep(3)
                    
                    if "Just a moment" in driver.title or "Cloudflare" in driver.title:
                        print(" -> Gặp Cloudflare, đang chờ tự động giải quyết...")
                        try:
                            WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.CSS_SELECTOR, "input[name='q']")))
                        except: pass
                        
                    # Gửi tìm kiếm qua JS để xuyên qua các popup quảng cáo/chặn
                    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, "input[name='q']")))
                    js_code = f"document.querySelector('input[name=\"q\"]').value = '{code}'; document.querySelector('form[action=\"/Search/\"]').submit();"
                    driver.execute_script(js_code)
                    
                    time.sleep(4)
                    if "Just a moment" in driver.title or "Cloudflare" in driver.title:
                        try:
                            WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.CSS_SELECTOR, "table.table-taxinfo")))
                        except: pass
                    
                    html = driver.page_source
                    extracted_data = extract_data_from_html(html, code)
                    
                    for key, value in extracted_data.items():
                        if key not in df.columns: df[key] = ""
                        df.at[index, key] = value
                    
                    company_name = extracted_data.get('Tên công ty', 'Không tìm thấy')
                    if company_name != "Không tìm thấy":
                        df.at[index, 'Trạng thái'] = "Thành công"
                        print(f" -> Hoàn tất: {company_name}")
                    else:
                        df.at[index, 'Trạng thái'] = "Không tìm thấy dữ liệu"
                        print(f" -> Lỗi: Không tìm thấy dữ liệu trên web")
                        
                    # Lưu tạm
                    if (index + 1) % 5 == 0:
                        try:
                            df.to_excel(output_file, index=False)
                            format_excel(output_file)
                        except PermissionError:
                            print(f" -> CẢNH BÁO: Đóng file {os.path.basename(output_file)} để hệ thống có thể lưu tạm!")
                            
                    time.sleep(3)
                except Exception as ex:
                    print(f" -> Lỗi trong quá trình cào: {ex}")

            print("\nĐã xong! Đang lưu file cuối cùng...")
            try:
                driver.quit()
            except: pass
            
            try:
                df.to_excel(output_file, index=False)
                format_excel(output_file)
                print(f"🎉 THÀNH CÔNG! Dữ liệu đã lưu vào: {os.path.basename(output_file)}")
                messagebox.showinfo("Thành công", f"Quá trình hoàn tất!\nFile kết quả: {os.path.basename(output_file)}")
                if self.var_auto_open.get():
                    try: os.startfile(output_file)
                    except: pass
            except PermissionError:
                backup = output_file.replace('.xlsx', f'_backup_{int(time.time())}.xlsx')
                df.to_excel(backup, index=False)
                format_excel(backup)
                print(f"⚠️ CẢNH BÁO: Đã lưu dự phòng sang {os.path.basename(backup)} vì file chính đang bị mở.")
                messagebox.showwarning("Cảnh báo", f"Đã lưu dự phòng sang {os.path.basename(backup)}\nvì file cũ đang mở.")
                if self.var_auto_open.get():
                    try: os.startfile(backup)
                    except: pass

        except Exception as e:
            print(f"LỖI NGHIÊM TRỌNG: {e}")
        finally:
            self.finish_processing()

    def finish_processing(self):
        self.is_running = False
        self.btn_run.config(state="normal", text="🚀 BẮT ĐẦU TRÍCH XUẤT")

if __name__ == "__main__":
    root = tk.Tk()
    app = MaSoThueApp(root)
    root.mainloop()
