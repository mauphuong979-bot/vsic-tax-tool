import sys
import os
import re
import time
import types
import pandas as pd
import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# ==========================================
# 1. POLYFILL (For Python 3.12+ compatibility)
# ==========================================
if 'distutils' not in sys.modules:
    class LooseVersion:
        def __init__(self, vstring):
            self.vstring = str(vstring)
            self.version = [int(x) if x.isdigit() else x for x in re.split(r'([0-9]+|\.)', self.vstring) if x and x != '.']
        def __str__(self): return self.vstring
        def __ge__(self, other): return True
        def __lt__(self, other): return False

    d = types.ModuleType('distutils')
    d.version = types.ModuleType('distutils.version')
    d.version.LooseVersion = LooseVersion
    sys.modules['distutils'] = d
    sys.modules['distutils.version'] = d.version

def format_excel(file_path):
    """Định dạng file Excel chuyên nghiệp"""
    try:
        time.sleep(1)
        wb = load_workbook(file_path)
        ws = wb.active
        
        default_font = Font(name='Times New Roman', size=11)
        header_font = Font(name='Times New Roman', bold=True, size=11)
        center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows():
            for cell in row:
                cell.font = default_font
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center", wrap_text=True)

        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = center_alignment
            cell.fill = PatternFill(fill_type=None)
            
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
                
        for idx, column in enumerate(ws.columns):
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        lines = str(cell.value).split('\n')
                        line_max = max([len(line) for line in lines])
                        if line_max > max_length:
                            max_length = line_max
                except: pass
            
            if idx < 4:
                adjusted_width = max(15, min(max_length + 5, 50))
            else:
                adjusted_width = max(10, min(max_length + 2, 60))
            ws.column_dimensions[column_letter].width = adjusted_width
            
        wb.save(file_path)
        return True
    except Exception as e:
        print(f"Error formatting Excel: {e}")
        return False

def get_chrome_main_version():
    """Tự động lấy phiên bản Chrome chính (Hỗ trợ cả Windows và Linux)"""
    # 1. Thử trên Windows
    try:
        import winreg
        key = winreg.OpenKey(winreg.HKEY_CURRENT_USER, r"Software\Google\Chrome\BLBeacon")
        version, _ = winreg.QueryValueEx(key, "version")
        return int(version.split('.')[0])
    except (ImportError, Exception):
        pass

    # 2. Thử trên Linux (Dành cho Cloud/VPS)
    try:
        import subprocess
        output = subprocess.check_output(['google-chrome', '--version'], stderr=subprocess.STDOUT)
        version = output.decode('utf-8').strip().split()[-1]
        return int(version.split('.')[0])
    except (FileNotFoundError, Exception):
        pass

    try:
        import subprocess
        output = subprocess.check_output(['chromium', '--version'], stderr=subprocess.STDOUT)
        version = output.decode('utf-8').strip().split()[-1]
        return int(version.split('.')[0])
    except (FileNotFoundError, Exception):
        pass

    return 125 # Giá trị mặc định an toàn nếu không tìm thấy

def setup_driver(headless=False, version_main=None):
    def get_options():
        options = uc.ChromeOptions()
        options.add_argument('--no-first-run --no-service-autorun --password-store=basic')
        options.add_argument('--no-sandbox')
        options.add_argument('--disable-dev-shm-usage')
        options.add_argument('--disable-gpu')
        if headless:
            options.add_argument('--headless')
        return options

    # Lượt 1: Thử với version_main
    try:
        driver = uc.Chrome(options=get_options(), version_main=version_main)
        return driver
    except Exception as e:
        print(f"Lỗi khởi tạo driver lượt 1 (với version {version_main}): {e}")
        
        # Lượt 2: Thử lại không có version_main (để uc tự nhận diện)
        try:
            return uc.Chrome(options=get_options())
        except Exception as e2:
            print(f"Lỗi khởi tạo driver lượt 2: {e2}")
            raise e2

def extract_data_from_html(html):
    soup = BeautifulSoup(html, 'html.parser')
    data = {}
    
    # 1. Ưu tiên lấy tên công ty từ các thẻ meta hoặc itemprop (thường không bị che bởi quảng cáo)
    company_name = ""
    name_tag = soup.find('th', itemprop='name') or soup.find('span', itemprop='name') or soup.find('h1', itemprop='name')
    
    if name_tag:
        company_name = name_tag.text.strip()
    
    # 2. Nếu không có, thử lấy từ tiêu đề trang (Title tag) - cấu trúc thường là: MST - Tên Công Ty - MaSoThue
    if not company_name or "Tra cứu mã số thuế" in company_name:
        title_tag = soup.find('title')
        if title_tag:
            title_text = title_tag.text.strip()
            # Regex để bóc tách: MST - TÊN CÔNG TY - MaSoThue
            match = re.search(r'^\d+\s*-\s*(.*?)\s*-\s*MaSoThue', title_text, re.IGNORECASE)
            if match:
                company_name = match.group(1).strip()
            elif " - " in title_text:
                parts = title_text.split(" - ")
                if len(parts) >= 2:
                    company_name = parts[1].strip()

    # 3. Fallback cuối cùng là H1
    if not company_name:
        h1 = soup.find('h1')
        if h1: company_name = h1.text.strip()

    data['Tên công ty'] = company_name if company_name else "Không tìm thấy"

    # Trích xuất bảng thông tin chính
    table = soup.find('table', class_='table-taxinfo')
    if table:
        for row in table.find_all('tr'):
            tds = row.find_all(['td', 'th'])
            if len(tds) == 2:
                label = tds[0].text.strip()
                value = tds[1].text.strip()
                # Xử lý riêng cho ngành nghề chính
                if 'Ngành nghề' in label or 'Mã ngành' in label:
                    label = 'Ngành nghề chính'
                    a_tag = tds[1].find('a')
                    if a_tag:
                        value = a_tag.text.strip()
                if label and label != company_name:
                    data[label] = value
                
    all_industries = []
    industry_tables = soup.find_all('table', class_='table')
    target_table = None
    
    for tbl in industry_tables:
        header = tbl.find('th')
        if header and ('Mã' in header.text or 'Ngành' in header.text):
            target_table = tbl
            break
            
    if not target_table and industry_tables:
        target_table = industry_tables[0]
        
    if target_table:
        for row in target_table.find_all('tr'):
            tds = row.find_all('td')
            if len(tds) >= 2:
                code = tds[0].text.strip()
                name = " ".join(tds[1].text.split())
                is_main = False
                if 'chính' in row.text.lower() or 'strong' in str(row):
                    is_main = True
                    if 'Ngành nghề chính' not in data or data['Ngành nghề chính'] == "":
                        data['Ngành nghề chính'] = f"{code} - {name}"
                    data['Mã Ngành Chính'] = code
                
                if is_main:
                    all_industries.append(f"[{code}] {name} (Ngành chính)")
                else:
                    all_industries.append(f"[{code}] {name}")
                    
    if all_industries:
        data['Danh sách ngành nghề'] = "\n".join(all_industries)
        
    if 'Mã Ngành Chính' not in data and 'Ngành nghề chính' in data:
        match = re.search(r'^\[?(\d{2,5})\]?[\s-]*', data['Ngành nghề chính'])
        if match:
            data['Mã Ngành Chính'] = match.group(1)
        else:
            data['Mã Ngành Chính'] = ""

    return data

class TaxScraper:
    def __init__(self, driver=None):
        self.driver = driver if driver else setup_driver()

    def search_tax_code(self, code):
        try:
            # Sử dụng URL tìm kiếm trực tiếp để bỏ qua bước nhập liệu thủ công
            search_url = f"https://masothue.com/Search/?q={code}&type=enterprise"
            self.driver.get(search_url)
            time.sleep(3)
            
            # Chờ trang chuyển hướng hoặc tải xong bảng dữ liệu
            # Tăng thời gian chờ để vượt qua các lớp bảo vệ/quảng cáo nếu có
            try:
                WebDriverWait(self.driver, 10).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "table.table-taxinfo"))
                )
            except:
                pass
            
            html = self.driver.page_source
            data = extract_data_from_html(html)
            
            # Kiểm tra nếu tên công ty lấy được là tên chung chung của trang web
            generic_titles = ["tra cứu mã số thuế", "mã số thuế", "masothue", "masothue.com"]
            company_name_lower = data.get('Tên công ty', '').lower()
            
            # Nếu tên công ty quá ngắn hoặc nằm trong danh sách từ khóa rác, và không có thông tin ngành nghề
            is_generic = any(title in company_name_lower for title in generic_titles)
            if (is_generic or len(company_name_lower) < 5) and 'Ngành nghề chính' not in data:
                return None
                
            return data
        except Exception as e:
            print(f"Error scraping {code}: {e}")
            return None

    def close(self):
        if self.driver:
            self.driver.quit()
